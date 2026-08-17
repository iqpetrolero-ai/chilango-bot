import os
import re
import unicodedata
from datetime import datetime, timezone, timedelta

import httpx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

import db
from menu import EMPAQUE

EXCEL_FILE = "pedidos_chilango.xlsx"
PERU_TZ = timezone(timedelta(hours=-5))
OWNER_PHONE = os.environ.get("OWNER_PHONE", "").strip()


def _normalizar(texto: str) -> str:
    """minúsculas, sin tildes, sin puntuación, sin 'de'/'del' — para comparar nombres de items."""
    texto = texto.lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\b(de|del)\b", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _split_top_level(items: str) -> list[str]:
    """Separa el texto de items por comas que NO estén dentro de paréntesis."""
    segmentos, actual, profundidad = [], [], 0
    for ch in items:
        if ch == "(":
            profundidad += 1
            actual.append(ch)
        elif ch == ")":
            profundidad -= 1
            actual.append(ch)
        elif ch == "," and profundidad == 0:
            segmentos.append("".join(actual))
            actual = []
        else:
            actual.append(ch)
    if actual:
        segmentos.append("".join(actual))
    return [s.strip() for s in segmentos if s.strip()]


def calcular_total_esperado(items: str) -> float | None:
    """Recalcula el total a partir de los precios reales del menú (BD).
    Devuelve None si no se pudo verificar con confianza (mejor no alertar que alertar mal)."""
    try:
        menu_items = db.get_menu_items()
    except Exception:
        return None
    precios = [(_normalizar(it["nombre"]), float(it["precio"])) for it in menu_items]
    if not precios:
        return None

    total = 0.0
    empaque_explicito = False
    algo_coincidio = False

    for seg in _split_top_level(items):
        seg_low = seg.lower()

        m_delivery = re.search(r"delivery\s*:?\s*s/\s*([\d.,]+)", seg_low)
        if m_delivery:
            total += float(m_delivery.group(1).replace(",", "."))
            algo_coincidio = True
            continue

        m_empaque = re.search(r"empaque\s*:?\s*s/\s*([\d.,]+)", seg_low)
        if m_empaque:
            total += float(m_empaque.group(1).replace(",", "."))
            empaque_explicito = True
            algo_coincidio = True
            continue

        m_paren = re.search(r"\(([^()]*)\)", seg)
        seg_sin_paren = re.sub(r"\([^()]*\)", "", seg).strip()
        if m_paren:
            paren_txt = m_paren.group(1).lower()
            if "extra" in paren_txt or "con " in paren_txt:
                # Ítem con extras/personalización dentro del paréntesis — precio ambiguo, no verificar.
                return None

        m_qty = re.match(r"^\s*(\d+)\s*x\s*(.+)$", seg_sin_paren, re.IGNORECASE)
        qty, nombre = (int(m_qty.group(1)), m_qty.group(2)) if m_qty else (1, seg_sin_paren)

        nombre_norm = _normalizar(nombre)
        if not nombre_norm:
            continue

        candidatos = [p for n, p in precios if n and (n in nombre_norm or nombre_norm in n)]
        if len(candidatos) != 1:
            return None  # sin match único → mejor no verificar que arriesgar falso positivo

        total += qty * candidatos[0]
        algo_coincidio = True

    if not algo_coincidio:
        return None
    if not empaque_explicito:
        total += EMPAQUE
    return round(total, 2)


def _extraer_monto(total_str: str) -> float | None:
    m = re.search(r"s/\s*([\d.,]+)", total_str.lower())
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


DELIVERY_MIN, DELIVERY_MAX = 4.00, 25.00  # rango típico visto en /admin/zonas-delivery


def verificar_total(items: str, total_str: str, direccion: str = "") -> str:
    """Devuelve un sufijo de alerta (para el WhatsApp del dueño) si el total del bot no
    coincide con el recalculado desde los precios reales del menú. Cadena vacía si todo
    bien o si no se pudo verificar con confianza.

    Si el pedido es delivery y el costo de envío NO está desglosado dentro de `items`
    (el bot a veces lo omite del tag aunque sí lo cobró), la diferencia observada podría
    ser un delivery legítimo no itemizado. En ese caso solo se alerta si la diferencia
    NO cae dentro del rango típico de costo de delivery (S/4–25) — una diferencia
    negativa, casi nula o absurdamente grande no se explica por un delivery faltante."""
    try:
        calculado = calcular_total_esperado(items)
        declarado = _extraer_monto(total_str)
        if calculado is None or declarado is None:
            return ""

        diff = declarado - calculado
        if abs(diff) <= 0.05:
            return ""

        es_recojo = direccion.strip().lower().startswith("recojo")
        tiene_delivery_en_items = bool(re.search(r"delivery", items, re.IGNORECASE))
        if not es_recojo and not tiene_delivery_en_items and DELIVERY_MIN <= diff <= DELIVERY_MAX:
            return ""  # posible delivery legítimo no itemizado — no se puede verificar con confianza

        return (
            f"\n\n⚠️ *VERIFICAR TOTAL* — el bot cobró S/ {declarado:.2f} pero "
            f"según los precios del menú (+ empaque{'' if es_recojo or tiene_delivery_en_items else ', sin contar delivery'}) "
            f"sería S/ {calculado:.2f}. Revisa el pedido."
        )
    except Exception as e:
        print(f"[VERIFICAR_TOTAL] Error: {e}")
    return ""


async def _send_whatsapp(to: str, body: str):
    """Envía un mensaje WhatsApp usando la API de Meta."""
    token = os.environ.get("META_ACCESS_TOKEN", "").strip()
    phone_number_id = os.environ.get("META_PHONE_NUMBER_ID", "").strip()
    to_clean = to.replace("+", "").replace(" ", "")
    if not token or not phone_number_id:
        print(f"[WA] ⚠️ META_ACCESS_TOKEN o META_PHONE_NUMBER_ID no configurados — no se envió WA a {to_clean}")
        return
    url = f"https://graph.facebook.com/v19.0/{phone_number_id}/messages"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    payload = {"messaging_product": "whatsapp", "to": to_clean, "type": "text", "text": {"body": body}}
    async with httpx.AsyncClient(timeout=10) as client:
        resp = await client.post(url, json=payload, headers=headers)
        if resp.status_code == 200:
            print(f"[WA] ✅ Enviado a {to_clean}")
        else:
            print(f"[WA] ❌ Error al enviar a {to_clean}: {resp.status_code} {resp.text}")


async def _send_telegram(chat_id: str, text: str):
    """Envía un mensaje por Telegram. Sin restricción de 24 horas."""
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    if not token or not chat_id:
        return False
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    async with httpx.AsyncClient(timeout=10) as client:
        resp = await client.post(url, json={
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "Markdown"
        })
        if resp.status_code == 200:
            print(f"[TELEGRAM] ✅ Enviado a chat_id {chat_id}")
            return True
        else:
            print(f"[TELEGRAM] ❌ Error a {chat_id}: {resp.status_code} {resp.text}")
            return False


async def _notify_delivery(delivery_phone: str, delivery_name: str,
                            delivery_index: int, mensaje_wa: str, mensaje_tg: str):
    """Envía notificación al motorizado por Telegram (preferido) o WhatsApp (fallback)."""
    tg_id = os.environ.get(f"DELIVERY_{delivery_index}_TELEGRAM_ID", "").strip()
    if tg_id:
        sent = await _send_telegram(tg_id, mensaje_tg)
        if sent:
            return  # Telegram OK — no necesita WA
    # Fallback a WhatsApp si no hay Telegram configurado o falló
    print(f"[DELIVERY] Usando WhatsApp para {delivery_name} (sin Telegram configurado)")
    await _send_whatsapp(delivery_phone, mensaje_wa)


DELIVERY_SERVICE_PHONE = os.environ.get("DELIVERY_SERVICE_PHONE", "").strip()
OWNER_PHONE = os.environ.get("OWNER_PHONE", "").strip()


async def notify_delivery_cost_query(phone_client: str, direccion: str,
                                      subtotal: str = "", items: str = "", pago: str = ""):
    """Notifica al dueño que un cliente quiere pagar delivery incluido.
    El dueño llama a Altoke para consultar el costo y lo ingresa en el panel."""
    msg_owner = (
        f"💰 *Cliente quiere pagar delivery incluido*\n"
        f"👤 +{phone_client}\n"
        f"📍 {direccion or 'Sin especificar'}\n"
        f"🛒 {items or '—'} · {subtotal or '—'}\n"
        f"📲 Llama a Altoke para el costo y regístralo en el panel."
    )
    await _send_whatsapp(OWNER_PHONE, msg_owner)

    # Guardar consulta pendiente para gestión desde el panel
    db.save_pending_cost_query(phone_client, subtotal, items, pago, direccion)
    print(f"[CONSULTAR_COSTO] ✅ Dueño notificado, consulta guardada para +{phone_client}")


def _init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedidos"

        headers = ["Fecha", "Hora", "Teléfono", "Items del Pedido", "Total", "Estado"]
        ws.append(headers)

        header_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12

        wb.save(EXCEL_FILE)


async def _notify_owner(phone_clean: str, items: str, total: str, metodo_pago: str, now: datetime, titulo: str = "🆕 *NUEVO PEDIDO — Chilango*", direccion: str = "", alerta: str = ""):
    try:
        token = os.environ.get("META_ACCESS_TOKEN", "").strip()
        phone_number_id = os.environ.get("META_PHONE_NUMBER_ID", "").strip()
        if not token or not phone_number_id:
            print("[NOTIFICACIÓN] META_ACCESS_TOKEN o META_PHONE_NUMBER_ID no configurados")
            return

        url = f"https://graph.facebook.com/v19.0/{phone_number_id}/messages"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        hora_str = now.strftime("%d/%m · %I:%M %p")

        template_name = os.environ.get("NOTIFY_TEMPLATE_NAME", "").strip()

        if template_name:
            # ── Modo template (permanente, sin restricción de 24 h) ──────────
            total_param = f"{total} ⚠️ VERIFICAR" if alerta else total
            payload = {
                "messaging_product": "whatsapp",
                "to": OWNER_PHONE,
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": "es"},
                    "components": [{
                        "type": "body",
                        "parameters": [
                            {"type": "text", "text": f"+{phone_clean}"},
                            {"type": "text", "text": items},
                            {"type": "text", "text": total_param},
                            {"type": "text", "text": metodo_pago},
                            {"type": "text", "text": hora_str},
                        ],
                    }],
                },
            }
            modo = "template"
        else:
            # ── Modo texto (requiere que el dueño haya escrito al bot hoy) ───
            pago_emoji = {"Yape/Plin": "💜 Yape/Plin", "Yape": "💜 Yape", "Plin": "💜 Plin", "Efectivo": "💵 Efectivo"}.get(metodo_pago, metodo_pago)
            dir_linea = f"\n📍 {direccion}" if direccion else ""
            mensaje = (
                f"{titulo}\n"
                f"👤 Cliente: +{phone_clean}\n"
                f"🛒 {items}\n"
                f"💰 {total}\n"
                f"💳 {pago_emoji}"
                f"{dir_linea}\n"
                f"🕒 {hora_str}"
                f"{alerta}"
            )
            payload = {
                "messaging_product": "whatsapp",
                "to": OWNER_PHONE,
                "type": "text",
                "text": {"body": mensaje},
            }
            modo = "texto"

        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(url, json=payload, headers=headers)
            if resp.status_code == 200:
                print(f"[NOTIFICACIÓN] ✅ Enviada al dueño ({OWNER_PHONE}) — modo {modo}")
            else:
                data = resp.json()
                error_msg = data.get("error", {}).get("message", resp.text)
                error_code = data.get("error", {}).get("code", resp.status_code)
                print(f"[ERROR NOTIFICACIÓN] Código {error_code}: {error_msg}")
                if error_code in (131047, 131026):
                    print("[ERROR NOTIFICACIÓN] El dueño no ha escrito al bot en las últimas 24h.")
                    print("[ERROR NOTIFICACIÓN] Solución: configura NOTIFY_TEMPLATE_NAME en Railway.")
    except Exception as e:
        print(f"[ERROR NOTIFICACIÓN] Excepción: {e}")


async def save_order(phone: str, items: str, total: str, metodo_pago: str = "Efectivo", direccion: str = "", notas: str = ""):
    now = datetime.now(PERU_TZ)
    phone_clean = phone.replace("whatsapp:", "").replace("+", "")

    # Persistencia confiable en SQLite
    db.save_order_db(phone_clean, items, total, metodo_pago, direccion, notas)
    print(f"[PEDIDO GUARDADO] {now.strftime('%d/%m %H:%M')} | {phone_clean} | {total} | {metodo_pago}")

    # Excel como backup (se pierde en reinicios sin Railway Volume)
    try:
        _init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        row = [
            now.strftime("%d/%m/%Y"),
            now.strftime("%H:%M"),
            phone_clean,
            items,
            total,
            "Nuevo 🆕",
            metodo_pago,
        ]
        ws.append(row)
        last_row = ws.max_row
        if last_row % 2 == 0:
            row_fill = PatternFill(start_color="F2F7EE", end_color="F2F7EE", fill_type="solid")
            for cell in ws[last_row]:
                cell.fill = row_fill
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[EXCEL] No se pudo guardar en Excel: {e}")

    alerta = verificar_total(items, total, direccion)
    if alerta:
        print(f"[VERIFICAR_TOTAL] ⚠️ Posible total incorrecto — {phone_clean} | {total}")
    await _notify_owner(phone_clean, items, total, metodo_pago, now, direccion=direccion, alerta=alerta)


async def update_order(phone: str, items: str, total: str, metodo_pago: str = "Efectivo", direccion: str = "", notas: str = ""):
    now = datetime.now(PERU_TZ)
    phone_clean = phone.replace("whatsapp:", "").replace("+", "")

    updated = db.update_latest_order(phone_clean, items, total, metodo_pago, direccion, notas)
    if updated:
        print(f"[PEDIDO MODIFICADO] {now.strftime('%d/%m %H:%M')} | {phone_clean} | {total} | {metodo_pago}")
        alerta = verificar_total(items, total, direccion)
        if alerta:
            print(f"[VERIFICAR_TOTAL] ⚠️ Posible total incorrecto (modificación) — {phone_clean} | {total}")
        await _notify_owner(
            phone_clean, items, total, metodo_pago, now,
            titulo="✏️ *PEDIDO MODIFICADO — Chilango*",
            direccion=direccion,
            alerta=alerta,
        )
    else:
        print(f"[PEDIDO MODIFICADO] No se encontró pedido activo para {phone_clean}")


async def cancel_order(phone: str):
    now = datetime.now(PERU_TZ)
    phone_clean = phone.replace("whatsapp:", "").replace("+", "")

    items_before = db.get_latest_active_order_items(phone_clean)
    cancelled = db.cancel_latest_order(phone_clean)
    if cancelled:
        if items_before and "gratis" in items_before.lower() and "pa ti solito" in items_before.lower():
            db.restore_promo_combo()
            print(f"[PROMO] 🔄 Combo gratis restaurado por cancelación — {phone_clean}")
        print(f"[PEDIDO CANCELADO] {now.strftime('%d/%m %H:%M')} | {phone_clean}")
        await _notify_owner(
            phone_clean, "—", "—", "—", now,
            titulo="❌ *PEDIDO CANCELADO — Chilango*",
        )
    else:
        print(f"[PEDIDO CANCELADO] No se encontró pedido activo para {phone_clean}")


def get_orders_count() -> int:
    return db.get_orders_count()

